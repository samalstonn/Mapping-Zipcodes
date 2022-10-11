import openpyxl
from excelconsts import *


def GardaBrinkstoNewSheet(store_info_workbook='storeinfotemplate.xlsx',
                          workbook1='gardatemplate.xlsx',
                          workbook2='brinkstemplate.xlsx'):
    """
    Creates a new sheet and appends store constants and minimum values
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    setStoreInfo(store_info_workbook, sheet)
    setHeaders(workbook1, sheet)
    setCurrentVendorInfo(sheet)

    listmins = findMinExRows(workbook1, workbook2)
    maxrows = listmins[-1]
    listmins.remove(listmins[-1])
    half = listmins.index('h')
    exvalueslist = listmins[:half]
    exvendorslist = listmins[half+1:]

    listmins = findMinIncRows(workbook2, workbook2)
    half = listmins.index('h')
    incvalueslist = listmins[:half]
    incvendorslist = listmins[half+1:]

    col = CHARGES_STARTING_COL
    x = 0
    for rows in range(2, maxrows):
        sheet.cell(row=rows, column=col, value=exvalueslist[x])
        sheet.cell(row=rows, column=col+1, value=exvendorslist[x])
        x += 1

    y = 0
    for rows in range(2, maxrows):
        sheet.cell(row=rows, column=col+2, value=incvalueslist[y])
        sheet.cell(row=rows, column=col+3, value=incvendorslist[y])
        y += 1
    setDimensions(sheet)

    wb.save('GardaBrinksCompliation.xlsx')


def setStoreInfo(store_info_workbook, sheet):
    """
    Sets the store information in the new sheet
    """
    wb = openpyxl.load_workbook(store_info_workbook)
    storesheet = wb.active
    maxrows = storesheet.max_row+1
    maxcols = storesheet.max_column+1
    list = []

    for r in range(1, maxrows):
        rows = []
        for c in range(1, maxcols):
            rows.append(storesheet.cell(row=r, column=c).value)
        list.append(rows)
    for r in list:
        sheet.append(r)


def setCurrentVendorInfo(sheet):
    """
    Sets Current Vendor and Current Pricing in the new sheet
    """


def setDimensions(sheet):
    """
    Sets each column's width to COLUMN_WIDTH in constants
    """
    for c in range(1, sheet.max_column+1):
        col = columnletter(c)
        sheet.column_dimensions[col].width = COLUMN_WIDTH


def lettercolumn(letter):
    """
    Letter to column number
    """
    letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p',
               'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
    return (letters.index(letter))+1


def columnletter(column):
    """"
    Column to Letter
    """
    letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p',
               'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
    return letters[column-1]


def findMinExRows(workbook1, workbook2=12):
    """
    Finds minimum of each column and orders them [min,vendor,min,vendor]
    """
    wb1 = openpyxl.load_workbook(workbook1)
    wb2 = openpyxl.load_workbook(workbook2)
    sheet1 = wb1.active
    sheet2 = wb2.active

    valueslist = []
    vendorslist = []

    for row in range(2, sheet1.max_row+1):
        first = sheet1.cell(row=row, column=GARDA_RATE_COLUMN).value
        second = sheet2.cell(row=row, column=BRINKS_RATE_COLUMN).value

        if not type(first) in [float, int] and first is not None:
            first = second[-6:]
            first = float(second)
        elif not type(second) in [float, int] and second is not None:
            second = second[-6:]
            second = float(second)
        try:
            mini = min(first, second)
            valueslist.append(mini)
            if mini == first:
                vendorslist.append('Garda')
            elif mini == second:
                vendorslist.append('Brinks')
        except:
            valueslist.append('No Bid')
            vendorslist.append('No Bid')

    vendorslist.append(sheet1.max_row+1)
    return valueslist + ['h'] + vendorslist


def findMinIncRows(workbook1, workbook2):
    """
    Finds minimum (taking into account fsc and security) of each column and
    orders them [min,vendor,min,vendor]
    """
    wb1 = openpyxl.load_workbook(workbook1)
    wb2 = openpyxl.load_workbook(workbook2)
    sheet1 = wb1.active
    sheet2 = wb2.active

    valueslist = []
    vendorslist = []

    for row in range(2, sheet1.max_row+1):
        first = sheet1.cell(row=row, column=GARDA_RATE_COLUMN).value
        second = sheet2.cell(row=row, column=BRINKS_RATE_COLUMN).value

        if not type(first) in [float, int] and first is not None:
            first = first[-6:]
            first = float(first)

        if not type(second) in [float, int] and second is not None:
            second = second[-6:]
            second = float(second)
        try:
            first = (first+(first*GARDA_FSC))+(first*GARDA_SECURITY)
            second = second+(second*BRINKS_FSC)
        except:
            pass

        try:
            mini = min(first, second)
            valueslist.append(mini)

            if mini == first:
                vendorslist.append('Garda')
            elif mini == second:
                vendorslist.append('Brinks')
        except:
            valueslist.append('No Bid')
            vendorslist.append('No Bid')

    return valueslist + ['h'] + vendorslist


def setHeaders(workbook, sheet):
    """
    Sets each header and original store information
    """
    wb = openpyxl.load_workbook(workbook)
    og_sheet = wb.active
    store_info = []
    x = 0
    maxrows = og_sheet.max_row

    cols = CHARGES_STARTING_COL
    sheet.cell(row=1, column=cols, value='Min Ex Surcharge')
    cols += 1
    sheet.cell(row=1, column=cols, value='Vendor')
    cols += 1
    sheet.cell(row=1, column=cols, value='Min Inc Surcharge')
    cols += 1
    sheet.cell(row=1, column=cols, value='Vendor')
    cols += 1
    sheet.cell(row=1, column=cols, value='Current Vendor Charge')
    cols += 1
    sheet.cell(row=1, column=cols, value='Vendor')


if __name__ == '__main__':
    print('Generating Sheet')
    GardaBrinkstoNewSheet()
    print('Sheet Generated')
